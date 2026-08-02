import csv
from pathlib import Path

from langchain_core.tools import tool

from config import GOOGLE_SHEET_LINK_PATH
from tools.common import get_logger, progress, retry

log = get_logger(__name__)

PLACEHOLDER_SPREADSHEET_IDS = {
    "your_spreadsheet_id",
    "{spreadsheet.id}",
    "spreadsheet_id",
}


@tool
@retry(max_attempts=1)
def verify_imports(csv_path: str, xlsx_path: str = "", spreadsheet_id: str = "") -> dict:
    """Verify that the Excel and/or Google Sheets imports actually succeeded.

    Re-reads the saved .xlsx file (if provided) and re-reads the Google
    Sheet (if a spreadsheet_id is provided) and compares their row counts
    against the source CSV. Call this LAST, after import_csv_to_excel and
    import_csv_to_google_sheets, to produce a trustworthy final report.

    Args:
        csv_path: path to the original source CSV.
        xlsx_path: path to the saved Excel workbook to verify (optional).
        spreadsheet_id: ID of the Google Sheet to verify (optional).

    Returns:
        dict with success flag and a per-target verification report.
    """
    src = Path(csv_path).expanduser().resolve()
    if not src.exists():
        return {"success": False, "error": f"CSV not found at {csv_path}"}

    with open(src, newline="", encoding="utf-8") as f:
        expected_rows = sum(1 for _ in csv.reader(f)) - 1

    report = {"expected_rows": expected_rows}

    if xlsx_path:
        progress("Verifying Excel workbook contents...")
        try:
            from openpyxl import load_workbook
            workbook_path = Path(xlsx_path).expanduser().resolve()
            wb = load_workbook(workbook_path, read_only=True)
            ws = wb.active
            actual = sum(1 for _ in ws.iter_rows(min_row=2)) if ws.max_row else 0
            report["excel"] = {
                "verified": actual == expected_rows,
                "rows_found": actual,
                "path": str(workbook_path),
            }
        except Exception as exc:  # noqa: BLE001
            report["excel"] = {"verified": False, "error": str(exc)}

    if spreadsheet_id:
        if spreadsheet_id.strip().lower() in PLACEHOLDER_SPREADSHEET_IDS:
            report["google_sheets"] = {
                "verified": False,
                "error": f"Invalid placeholder spreadsheet_id received: {spreadsheet_id}",
            }
            report["success"] = False
            progress("Verification skipped for Google Sheets because spreadsheet_id was a placeholder")
            return report

        progress("Verifying Google Sheet contents...")
        try:
            from tools.gsheets_tool import _get_client
            client = _get_client()
            sh = client.open_by_key(spreadsheet_id)
            values = sh.sheet1.get_all_values()
            actual = max(len(values) - 1, 0)
            url = sh.url or f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
            report["google_sheets"] = {
                "verified": actual == expected_rows,
                "rows_found": actual,
                "url": url,
            }
            progress(f"Verified Google Sheet URL: {url}")
            log.info("Verified Google Sheet URL: %s", url)
            GOOGLE_SHEET_LINK_PATH.parent.mkdir(parents=True, exist_ok=True)
            GOOGLE_SHEET_LINK_PATH.write_text(url + "\n", encoding="utf-8")
            progress(f"Saved verified Google Sheet link to {GOOGLE_SHEET_LINK_PATH}")
        except Exception as exc:  # noqa: BLE001
            report["google_sheets"] = {"verified": False, "error": str(exc)}

    overall = all(
        report.get(k, {}).get("verified", True)
        for k in ("excel", "google_sheets")
        if k in report
    )
    report["success"] = overall
    progress(f"Verification complete - overall success: {overall}")
    return report
